import time
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import TemplateView, ListView, CreateView
from django.contrib import messages
from django.core.paginator import Paginator
from datetime import datetime
from django.utils.timezone import make_aware
import openpyxl
from django.core.files.storage import default_storage
from core.utils.excell_task import upload_func



from core.forms import AddUserForm, UploadExcelForm
from core.models import Account, ExcellAsanInfo, ExcellDeclInfo
# from core.tasks import assign_order, login_to_asan, scrape_and_save_packages
from delivery.models import Package



@login_required
def create_account(request):
    if request.method == 'POST':
        form = AddUserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = AddUserForm()
    return render(request, 'account/add_user.html', {'form': form})


class IndexView(LoginRequiredMixin, ListView):
    template_name = 'home/accounts.html'
    model = Account
    context_object_name = 'accounts'

    def get_queryset(self):
        return Account.objects.filter(is_processed=True, done=False).order_by('-processed_at')


def add_account_to_process(request):
    client_id = request.POST.get('client_id')

    account = Account.objects.get(client_id=client_id)

    account.is_processing = True
    account.is_processed = True
    account.processed_at = make_aware(datetime.now())
    account.save()
    try:
        messages.success(request, f"{account} ID-li istifadəçi üçün prosesə başlanıldı")
        scrape_and_save_packages.apply_async(args=[account.id])
    except Exception as e:
        account.is_processing = False
        account.save()
        messages.error(request, f"{account} ID-li istifadəçidə problem yarandı, giriş məlumatlarını təkrar yoxlayın")
        print(e)

    accounts = Account.objects.filter(is_processed=True, done=False).order_by('-updated_at')
    return render(request, 'partials/account-list.html', {'accounts': accounts})


def search_account(request):
    search_text = request.POST.get('search')
    results = Account.objects.filter(client_id__icontains=search_text).exclude(is_processing=True)
    context = {"results": results}
    return render(request, 'partials/search-results.html', context)


def clear(request):
    return HttpResponse("")


@login_required
def account_detail(request, client_id_slug):
    account = Account.objects.get(slug=client_id_slug)
    packages = account.packages.filter(done=False).order_by('created_at')
    page_number = request.GET.get('page', 1)
    paginator = Paginator(packages, 9)
    page_obj = paginator.get_page(page_number)
    context = {
        "results": page_obj,
        'account': account
    }
    if request.htmx:
        return render(request, 'account/account_packages_elements.html', context)
    return render(request, 'account/account_packages.html', context)


def done_package(request, package_id):
    package = Package.objects.get(id=package_id)
    package.done = True
    package.save()
    return HttpResponse(status=200)


def upload_asan_excell(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('upload-excell-asan-login')
    else:
        form = UploadExcelForm()

    return render(request, 'excell/upload1.html', {'form': form})


def upload_user_id_and_shipingId(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('upload-excell-user-info')
    else:
        form = UploadExcelForm()

    return render(request, 'excell/upload2.html', {'form': form})




# =============================MASI===================================
def process_excell_user_info(request):
    
    record1 = ExcellAsanInfo.objects.last()
    record2 = ExcellDeclInfo.objects.last()

    return HttpResponse(upload_func(record1, record2))


# --------------------------------------------------------------------------------


list_ = []
def process_excell_decl_info(request):

    import openpyxl
    from django.core.files.storage import default_storage

    # must be append to list excel data

    # Get the last record
    record = ExcellDeclInfo.objects.last()

    if record and record.file:
        print('record.file.name', record.file.name)
        # Open the Excel file using openpyxl
        file_path = default_storage.path(record.file.name)
        print('file_path', file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Assuming the headers are in the first row
        # headers = [cell.value for cell in next(sheet.iter_rows(values_only=True))]
        headers = [cell for cell in next(sheet.iter_rows(values_only=True))]

        # Verify if headers are present
        if all(header in headers for header in ["ShipmentID", "UserID"]):
            for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row
                list_.append({
                    "ShipmentID": row[headers.index("ShipmentID")],
                    "UserID": row[headers.index("UserID")],
                })

        workbook.close()
        # for i in list_:
        #     assign_order(i.get("UserID"), i.get("ShipmentID"))
        #     time.sleep(50)
        # print("list_", list_)            
        return HttpResponse(list_)