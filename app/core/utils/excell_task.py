
import random
from django.http import HttpResponse
from django.utils.timezone import make_aware
import openpyxl
from django.core.files.storage import default_storage


# ------------------------------------------------------------
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.support.ui import WebDriverWait

from core.models import Declaration, FailedDeclar, Interval


chrome_driver_path = "/home/safex/chromedriver"

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36"


def login_to_asan(DecID, FIN, PassWord, UserID):
    try:
        print("======================  başladıq =========================")
        options = webdriver.ChromeOptions()
        options.add_argument("--no-sandbox")
        # options.add_argument('--headless')
        options.add_argument("--disable-dev-shm-usage")
        ser = Service(executable_path="/home/taleh/Downloads/chromedriver/chromedriver")
        driver = webdriver.Chrome(options=options, service=ser)

        driver.get(
            "https://asanlogin.my.gov.az/auth?origin=https:%2F%2Fe.customs.gov.az%2Fauth%2Fasan"
        )
        # driver.maximize_window()
        time.sleep(4)
        print("======================  asan loginleshdi =========================")
        # must be click tag which it inneHTML is equal İdentifikasiya nömrəsi ilə
        # //*[@id="wrapper"]/div/div/div/div/app-login-tabs/div/div[1]/div[1]/div/p[1]
        element_p_tag = driver.find_element(
            By.XPATH,
            '//*[@id="wrapper"]/div/div/div/div/app-login-tabs/div/div[1]/div[1]/div/p[1]',
        )
        element_p_tag.click()
        time.sleep(4)
        # login 6ARPY3L
        # password Taleh123
        element1_login_input = driver.find_element(
            By.XPATH, '//*[@id="loginform"]/div[1]/div/div/input'
        )
        element1_password_input = driver.find_element(
            By.XPATH, '//*[@id="loginform"]/div[2]/div/input'
        )
        element1_login_input.send_keys(FIN)
        element1_password_input.send_keys(PassWord)
        time.sleep(2)

        # get send button
        element1_send_button = driver.find_element(
            By.XPATH, '//*[@id="wrapper"]/div/div/div/div/app-log-in/div[3]/div/div[2]/div'
        )
        element1_send_button.click()
        time.sleep(4)

        menu_button = driver.find_element(By.CLASS_NAME, "jss67")
        menu_button.click()
        time.sleep(4)

        ul_tag1 = driver.find_element(By.CLASS_NAME, "jss197")
        time.sleep(2)

        sec_2 = ul_tag1.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[3]/ul/li[2]/ul/li[1]')
        sec_2.click()
        time.sleep(2)

        post_dec = driver.find_element(
            By.XPATH,
            '//*[@id="root"]/div[1]/div[1]/div[2]/div/div/section/div/div[2]/div/div[2]',
        )

        post_dec.click()
        # driver.get("https://e.customs.gov.az/for-individuals")
        time.sleep(4)


        div_items_x_path = '/html/body/div[4]/div[3]/div[2]'
        result = driver.find_element(By.XPATH, div_items_x_path)
        items = result.find_elements(By.CLASS_NAME, "MuiPaper-root")
        time.sleep(2)
        for index, item in enumerate(items):
            track_span = item.find_element(By.CLASS_NAME, "MuiTypography-body1").text
            track_number = track_span.split(":")[1].strip()
            if DecID == track_number:
                if index != 0:
                    svg_icon = item.find_element(By.TAG_NAME, "svg")
                    svg_icon.click()
                    time.sleep(2)
                body_ul = item.find_element(By.CLASS_NAME, "MuiList-dense")
                body_ul_items = body_ul.find_elements(By.TAG_NAME, "li")
                print("invoys", body_ul_items[2].text)
                invoys_price = body_ul_items[2].text

                time.sleep(4)
                confirm_button = item.find_element(
                    By.CLASS_NAME, "MuiButton-containedPrimary"
                )
                confirm_button.click()
                time.sleep(4)

                next_button2 = driver.find_element(
                    By.CLASS_NAME, "MuiButton-containedPrimary"
                )
                next_button2.click()
                time.sleep(4)

                next_button3 = driver.find_element(
                    By.CLASS_NAME, "MuiButton-containedPrimary"
                )
                next_button3.click()

                time.sleep(4)
                next_button4 = driver.find_element(By.CLASS_NAME, "MuiButton-contained")
                next_button4.click()

                time.sleep(4)
                form_tag = driver.find_element(By.TAG_NAME, "form")
                form_inputs = form_tag.find_elements(
                    By.CLASS_NAME, "MuiOutlinedInput-input"
                )
                popups = form_tag.find_elements(
                    By.CLASS_NAME, "MuiOutlinedInput-inputAdornedEnd"
                )
                first_popup = popups[0]
                first_popup.send_keys("Paltar")
                paltar_option = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//*[contains(text(), 'Paltar')]")
                    )
                )
                paltar_option.click()
                time.sleep(2)

                second_popup = popups[1]
                second_popup.send_keys("Digər")
                diger_option = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//*[contains(text(), 'Digər')]")
                    )
                )
                diger_option.click()
                time.sleep(2)


                third_popup = popups[2]
                third_popup.send_keys("Türk lirəsi")
                turk_lirasi_option = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//*[contains(text(), 'Türk lirəsi')]")
                    )
                )
                turk_lirasi_option.click()

                time.sleep(2)
                invoys_price_input = form_inputs[2]

                delivery_cost = form_inputs[3]
                qs = Interval.objects.filter(is_active=True).order_by("price")
                for i in qs:    
                    if delivery_cost > i.price:
                        print("i.price", i.price)
                        print("i.start_interval", i.start_interval)
                        print("i.end_interval", i.end_interval)
                        print('delivery_cost', delivery_cost)
                        random_pr_price = random.randint(i.start_interval, i.end_interval)
                        invoys_price_input.send_keys(random_pr_price)
                        break
                
                # kommente aldim deyesen artiqdi
                # invoys_price_input = form_inputs[2]
                # invoys_price_input.send_keys(invoys_price)

                # invoys_price_input = form_inputs[2]
                # invoys_price_input.send_keys(invoys_price)
                time.sleep(4)

                quantity_select_input = form_inputs[5]
                quantity_select_input.send_keys("1")
                time.sleep(4)

                next_button5 = form_tag.find_element(
                    By.CLASS_NAME, "MuiButton-containedPrimary"
                )
                next_button5.click()

                time.sleep(4)
                # class jss445
                checkbox_input = driver.find_element(By.CLASS_NAME, "jss445")
                checkbox_input.click()

                time.sleep(2)
                confirm_button = driver.find_element(By.CSS_SELECTOR, 'button[test-id="approve"]')
                confirm_button.click()

                time.sleep(3)
                failed_dec = FailedDeclar.objects.filter(fin_code=FIN, password=PassWord, user_id=UserID, dec_id=DecID, is_active=True).first()
                if failed_dec:
                    failed_dec.is_active = False
                    failed_dec.save()
                declaration = Declaration.objects.filter(fin_code=FIN, password=PassWord, user_id=UserID, dec_id=DecID).first()
                if declaration:
                    declaration.is_declared = True
                    declaration.save()
            driver.quit()
    except Exception as e:
        FailedDeclar.objects.create(fin_code=FIN, password=PassWord, user_id=UserID, dec_id=DecID, reason=e)



# ----------------------------------------------------------------  




def upload_func(record1, record2):
    list_1 = [] 
    list_2 = []
    

    if record1 and record1.file:
        # Open the Excel file using openpyxl
        file_path = default_storage.path(record1.file.name)
        # print("file_path", file_path)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Assuming the headers are in the first row
        # headers = [cell.value for cell in next(sheet.iter_rows(values_only=True))]
        headers = [cell for cell in next(sheet.iter_rows(values_only=True))]

        # Verify if headers are present
        if all(header in headers for header in ["UserID", "FIN", "Pass"]):
            for row in sheet.iter_rows(min_row=2, values_only=True):
                list_1.append(
                    {
                        "UserID": row[headers.index("UserID")],
                        "FIN": row[headers.index("FIN")],
                        "Pass": row[headers.index("Pass")] if len(str(row[headers.index("Pass")]).split()) == 1 else str(row[headers.index("Pass")]).split()[0],
                    }
                )
        # print("list_1", list_1)
        workbook.close()


    if record2 and record2.file:
        # Open the Excel file using openpyxl
        file_path = default_storage.path(record2.file.name)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Assuming the headers are in the first row
        # headers = [cell.value for cell in next(sheet.iter_rows(values_only=True))]
        headers = [cell for cell in next(sheet.iter_rows(values_only=True))]

        # Verify if headers are present
        if all(header in headers for header in ["ShipmentID", "UserID"]):
            for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row
                list_2.append({
                    "ShipmentID": row[headers.index("ShipmentID")],
                    "UserID": row[headers.index("UserID")],
                })

        workbook.close()
        
        for user in list_1:
            for dec in list_2:
                if user["UserID"] == dec["UserID"]:
                    time.sleep(5)
                    print('dec["ShipmentID"]', dec["ShipmentID"])
                    print('user["FIN"],', user["FIN"],)
                    print('user["Pass"]', user["Pass"])
                    login_to_asan(dec["ShipmentID"], user["FIN"], user["Pass"], dec["UserID"])

        return  HttpResponse("masiltilar yuklendi")
